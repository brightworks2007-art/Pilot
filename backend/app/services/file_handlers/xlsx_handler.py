import io

from openpyxl import Workbook, load_workbook

from app.services.file_handlers.base import FileHandler


class XlsxHandler(FileHandler):
    """Handles .xlsx via openpyxl. Content is treated as CSV-like rows."""

    def read_text(self, file_bytes: bytes) -> str:
        wb = load_workbook(io.BytesIO(file_bytes))
        sheet = wb.active
        lines = []
        for row in sheet.iter_rows(values_only=True):
            lines.append(",".join("" if v is None else str(v) for v in row))
        return "\n".join(lines)

    def create(self, content_text: str) -> bytes:
        wb = Workbook()
        sheet = wb.active
        for line in content_text.split("\n"):
            sheet.append(line.split(","))
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def update(self, file_bytes: bytes, instruction_result_text: str) -> bytes:
        wb = Workbook()
        sheet = wb.active
        for line in instruction_result_text.split("\n"):
            sheet.append(line.split(","))
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
